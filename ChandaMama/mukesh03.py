import google.generativeai as genai
import speech_recognition as sr
import pyttsx3
import requests
import pygame
import json
import re
from io import BytesIO
import pandas as pd
import openpyxl


# Configure 11Labs API key
genai.configure(api_key="<INSERT_LLM_API_KEY_HERE>")

# Initialize text-to-speech engine
engine = pyttsx3.init()

# Initialize GenerativeModel
model = genai.GenerativeModel(model_name="gemini-1.0-pro")

# Function to speak text using 11Labs API
def say_with_11labs(text):
    url = "https://api.elevenlabs.io/v1/text-to-speech/<VOICE_ID>"
    headers = {
        "Accept": "audio/mpeg",
        "Content-Type": "application/json",
        "xi-api-key": "<INSERT_11LAB_API_KEY_HERE>"
    }
    data = {
        "text": text,
        "model_id": "eleven_monolingual_v1",
        "voice_settings": {
            "stability": 0.5,
            "similarity_boost": 0.5,
            "voice_id": "<VOICE_ID>",
        }
    }
    response = requests.post(url, json=data, headers=headers)
    if response.status_code == 200:
        pygame.mixer.init()
        pygame.mixer.music.load(BytesIO(response.content))
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            pygame.time.Clock().tick(10)
        print("Audio playback complete.")
    else:
        print("Error occurred while processing text-to-speech:", response.status_code)
        print("Response content:", response.text)

# Replace the original 'say' function with 'say_with_11labs'
say = say_with_11labs

# Set up the model
generation_config = {
    "temperature": 0.6,
    "top_p": 1,
    "top_k": 1,
    "max_output_tokens": 2048,
}

# Function to capture voice command
def takeCommand():
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        recognizer.adjust_for_ambient_noise(source)
        print("Listening...")
        audio = recognizer.listen(source)

        try:
            query = recognizer.recognize_google(audio, language="en-in")
            print(f"User: {query}")
            return query.lower()
        except sr.UnknownValueError:
            return ""
        except sr.RequestError:
            say("Sorry, I couldn't understand that.")
            return ""

safety_settings = [
    {
        "category": "HARM_CATEGORY_HARASSMENT",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE"
    },
    {
        "category": "HARM_CATEGORY_HATE_SPEECH",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE"
    },
    {
        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE"
    },
    {
        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE"
    },
]

model = genai.GenerativeModel(model_name="gemini-1.0-pro",
                              generation_config=generation_config,
                              safety_settings=safety_settings)

convo = model.start_chat(history=[
    {
        "role": "user",
        "parts": ["You are a conversational prompt for a Customer Virtual Assistant (CVA) named Scogo, designed to accurately interpret and gather essential information from customers. Your prompt should guide the interaction step by step, asking one piece of information at a time like dont ask two question at once, ensuring the caller's information is collected effectively. As the user, I'll provide inputs, and you'll wait for my response after every question.To begin, you'll greet me with \"Hi\" and wait for my response. After my response, you'll ask about the preferred language for the conversation, offering English or Hindi as options. If i select english continue the whole converation in english and greet the user. For example, if English is chosen, you'll say, \"Hi, this is Scogo, your smart solutions assistant. How are you?\" .If i select Hindi you will continue the conversation in hindi but using english sentences so the text can be analysed further. If Hindi is chosen, for example \"Namaste, me hu Scogo, me aapki kese sahayata kar sakta hu?\".  i will then greet you handle my response to your greeting accordingly, after that ask my issue and then if it is for first time or existing issue. Question me about all the information full name of the caller, the nature of the issue (new or existing), contact information(check if it is ten digits if not then ask me to say it again), address with pin code, details of the problem, then ask the product name and its model number(model number is optional if user does not know the model number) one by one only don't ask multiple question in same response and at the end ask the user to confirm a technician visit or call, the date and time should be on working days and between day time, if the user response with other than this time suggest then a relevant time after getting all the required info and confirm the visit, thank the user for calling and say goodbye. With this good bye which is last response also print code for .json file which is your last response consisting of all information which is full name of the caller, the nature of the issue (new or existing), contact information, address with pin code, details of the problem, product name, model number and preferred date and time for technician visits with variable names (user_name, user_issue, user_contact, user_address, user_pin_code,issue_details,product_name,model_number, visit_date,visit_time). the format should be\n{\n\"user_name\": \"\",\n\"user_issue\": \"\",\n\"user_contact\": \"\",\n\"user_address\": \"\",\n\"user_pin_code\": \"\",\n\"issue_details\": \"\",\n\"product_name\": \"\",\n\"model_number\": \"\",\n\"visit_date\": \"\",\n\"visit_time\": \"\"\n}\n\nLet's start your response by saying just \"Hi\" and wait for my response"]
    },
    {
        "role": "model",
        "parts": ["Hi"]
    },
])

# Flag to check if JSON data extraction is in progress
extracting_json = False
json_data = ""

# Regular expression pattern to match JSON object
pattern = r'\{[^{}]+\}'


while True:
    user_input = takeCommand()

    # Check if user input is empty
    if user_input:
        # Send user input to the conversation
        convo.send_message(user_input)

        # Get response from the conversation
        print(convo.last.text)
        scogo_response = convo.last.text

        # Speak the response
        say(scogo_response)

        # Check for goodbye and terminate
        if any(keyword in user_input.lower() for keyword in ["goodbye", "bye", "see you later", "chaliye rakhta hu", "rakhu phir?", "alvida"]):
            break

    # Find JSON object in the response
    match = re.search(pattern, scogo_response)

    if match:
        # Extract JSON object
        json_data = match.group()

        # Parse JSON data
        data = json.loads(json_data)

        # Print extracted information
        print("Ticket Raised Successfully!:")
        # print(json.dumps(data, indent=2))
        json_file_path = "Ticket.json"
        with open(json_file_path, "w") as json_file:
            json.dump(data, json_file)
    else:
        print("")


# Create a DataFrame from the dictionary
df = pd.DataFrame(data, index=[0])  # Add an index (optional)

# Load the existing Excel file or create a new one
try:
    workbook = openpyxl.load_workbook("user_data.xlsx")  # Load existing file
    worksheet = workbook.active
    next_row = worksheet.max_row + 1  # Find the next available row
except FileNotFoundError:
    workbook = openpyxl.Workbook()  # Create a new file if it doesn't exist
    worksheet = workbook.active
    next_row = 2  # Start writing from row 2 (leaving row 1 for header)

# Write column headers if the file is new
if next_row == 2:
    for col in df.columns:
        cell = worksheet.cell(row=1, column=df.columns.get_loc(col) + 1)
        cell.value = col

# Write data to the next available row
for col_num in range(1, df.shape[1] + 1):
    cell = worksheet.cell(row=next_row, column=col_num)
    cell.value = df.iloc[0, col_num - 1]

workbook.save("C:/Users/Durvesh/Desktop/ChandaMama/user_Data.xlsx")

# Save the Excel file
# workbook.save("user_data.xlsx")